from flask import Flask, request, jsonify, session, send_from_directory
from flask_cors import CORS
import json, os, uuid, random, string
from datetime import datetime, timedelta
import openpyxl
from io import BytesIO
from flask import send_file

app = Flask(__name__, static_folder='.')
app.secret_key = 'streamvault_secret_2024'
CORS(app)

DB_FILE = 'database.json'

# ─── DB helpers ───────────────────────────────────────────────────────────────
def load_db():
    if not os.path.exists(DB_FILE):
        return default_db()
    with open(DB_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_db(db):
    with open(DB_FILE, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=2)

def default_db():
    return {
        "platforms": [
            {"id":"netflix","name":"Netflix","category":"streaming","image":"https://upload.wikimedia.org/wikipedia/commons/0/08/Netflix_2015_logo.svg","description":"Películas, series y documentales en 4K. Perfil compartido con PIN exclusivo.","delivery":"perfil","stock":5,"prices":{"30":15,"60":25,"90":35},"active":True},
            {"id":"prime","name":"Prime Video","category":"streaming","image":"https://upload.wikimedia.org/wikipedia/commons/1/11/Amazon_Prime_Video_logo.svg","description":"Series y películas Amazon Originals. Acceso por perfil individual.","delivery":"perfil","stock":8,"prices":{"30":10,"60":18,"90":25},"active":True},
            {"id":"hbo","name":"HBO Max","category":"streaming","image":"https://upload.wikimedia.org/wikipedia/commons/1/17/HBO_Max_Logo.svg","description":"Series HBO, DC y Warner Bros. Perfil con PIN protegido.","delivery":"perfil","stock":6,"prices":{"30":12,"60":20,"90":30},"active":True},
            {"id":"disney","name":"Disney+","category":"streaming","image":"https://upload.wikimedia.org/wikipedia/commons/3/3e/Disney%2B_logo.svg","description":"Marvel, Star Wars, Pixar y National Geographic. Perfil familiar.","delivery":"perfil","stock":7,"prices":{"30":10,"60":18,"90":26},"active":True},
            {"id":"disneypremium","name":"Disney Premium","category":"streaming","image":"https://upload.wikimedia.org/wikipedia/commons/3/3e/Disney%2B_logo.svg","description":"Acceso completo Disney+ con cuenta exclusiva sin compartir.","delivery":"cuenta","stock":3,"prices":{"30":20,"60":35,"90":50},"active":True},
            {"id":"vix","name":"VIX Premium","category":"streaming","image":"https://upload.wikimedia.org/wikipedia/commons/thumb/b/bc/VIX_logo.svg/1200px-VIX_logo.svg.png","description":"Cine y series en español. Deportes y telenovelas premium.","delivery":"perfil","stock":4,"prices":{"30":8,"60":14,"90":20},"active":True},
            {"id":"crunchyroll","name":"Crunchyroll","category":"streaming","image":"https://upload.wikimedia.org/wikipedia/commons/0/08/Crunchyroll_Logo.svg","description":"Anime en simulcast, doblado y subtitulado. La mayor biblioteca anime.","delivery":"perfil","stock":6,"prices":{"30":8,"60":14,"90":20},"active":True},
            {"id":"spotify","name":"Spotify Premium","category":"musica","image":"https://upload.wikimedia.org/wikipedia/commons/2/26/Spotify_logo_with_text.svg","description":"Música sin anuncios, descargas offline. Activación a tu correo.","delivery":"correo","stock":10,"prices":{"30":8,"60":14,"90":20},"active":True},
            {"id":"youtube","name":"YouTube Premium","category":"musica","image":"https://upload.wikimedia.org/wikipedia/commons/b/b8/YouTube_Logo_2017.svg","description":"Sin anuncios en YouTube y YouTube Music. Activación por correo.","delivery":"correo","stock":8,"prices":{"30":10,"60":18,"90":25},"active":True},
            {"id":"tidal","name":"Tidal HiFi","category":"musica","image":"https://upload.wikimedia.org/wikipedia/commons/thumb/8/84/Tidal_logo.svg/1200px-Tidal_logo.svg.png","description":"Audio en máxima calidad HiFi y MQA. Exclusivas de artistas.","delivery":"correo","stock":5,"prices":{"30":12,"60":20,"90":28},"active":True},
            {"id":"deezer","name":"Deezer Premium","category":"musica","image":"https://upload.wikimedia.org/wikipedia/commons/e/e3/Deezer_logo.svg","description":"90 millones de canciones en alta calidad. Activación por correo.","delivery":"correo","stock":6,"prices":{"30":8,"60":14,"90":20},"active":True},
            {"id":"canva","name":"Canva Pro","category":"musica","image":"https://upload.wikimedia.org/wikipedia/commons/b/bb/Canva_Logo.svg","description":"Diseño profesional con plantillas premium, fondo removedor y más. Invitación a tu correo. 1 año.","delivery":"correo_manual","stock":15,"prices":{"365":25},"active":True},
            {"id":"gemini","name":"Gemini AI","category":"ia","image":"https://www.gstatic.com/lamda/images/gemini_sparkle_v002_d4735304ff6292a690345.svg","description":"IA de Google con capacidades multimodales avanzadas. Por dispositivo/correo.","delivery":"dispositivo","stock":5,"prices":{"30":15,"60":25,"90":35},"active":True},
            {"id":"chatgpt","name":"ChatGPT Plus","category":"ia","image":"https://upload.wikimedia.org/wikipedia/commons/0/04/ChatGPT_logo.svg","description":"GPT-4o con plugins, DALL-E y análisis avanzado. Por correo personal.","delivery":"dispositivo","stock":4,"prices":{"30":18,"60":30,"90":42},"active":True},
            {"id":"perplexity","name":"Perplexity Pro","category":"ia","image":"https://upload.wikimedia.org/wikipedia/commons/thumb/1/1d/Perplexity_AI_logo.svg/1200px-Perplexity_AI_logo.svg.png","description":"Búsqueda con IA en tiempo real, modelos avanzados y sin límites.","delivery":"dispositivo","stock":6,"prices":{"30":12,"60":20,"90":28},"active":True},
            {"id":"claude","name":"Claude AI Pro","category":"ia","image":"https://upload.wikimedia.org/wikipedia/commons/thumb/8/8a/Claude_AI_logo.svg/1200px-Claude_AI_logo.svg.png","description":"Claude Pro de Anthropic con uso prioritario y acceso a todos los modelos.","delivery":"dispositivo","stock":5,"prices":{"30":15,"60":25,"90":35},"active":True}
        ],
        "accounts": [],
        "orders": [],
        "clients": [],
        "finances": {"ingresos": [], "egresos": []},
        "payment_codes": {},
        "admin_password": "admin2024"
    }

# ─── Static files ──────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/admin')
def admin():
    return send_from_directory('.', 'admin.html')

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('.', path)

# ─── Public API ───────────────────────────────────────────────────────────────
@app.route('/api/platforms', methods=['GET'])
def get_platforms():
    db = load_db()
    return jsonify([p for p in db['platforms'] if p.get('active')])

@app.route('/api/verify-payment', methods=['POST'])
def verify_payment():
    data = request.json
    code = data.get('code', '').strip()
    method = data.get('method', 'yape')
    platform_id = data.get('platform_id')
    days = str(data.get('days'))
    client_name = data.get('client_name', '')
    client_whatsapp = data.get('client_whatsapp', '')
    client_email = data.get('client_email', '')
    
    db = load_db()
    platform = next((p for p in db['platforms'] if p['id'] == platform_id), None)
    
    if not platform:
        return jsonify({"success": False, "message": "Plataforma no encontrada"})
    
    if platform.get('stock', 0) <= 0:
        return jsonify({"success": False, "message": "Sin stock disponible"})

    # For Yape: verify 3-digit code
    if method == 'yape':
        if not code or len(code) != 3 or not code.isdigit():
            return jsonify({"success": False, "message": "Código Yape inválido. Debe ser 3 dígitos."})
        
        # Check if code already used
        if code in db.get('payment_codes', {}):
            return jsonify({"success": False, "message": "Este código ya fue utilizado anteriormente."})
        
        # Mark code as used
        db['payment_codes'][code] = {"used_at": datetime.now().isoformat(), "platform": platform_id}
    
    # For Canva - always manual
    if platform['delivery'] == 'correo_manual':
        order_id = str(uuid.uuid4())[:8].upper()
        order = {
            "id": order_id,
            "platform_id": platform_id,
            "platform_name": platform['name'],
            "client_name": client_name,
            "client_whatsapp": client_whatsapp,
            "client_email": client_email,
            "days": days,
            "price": platform['prices'].get(days, 0),
            "payment_method": method,
            "payment_code": code,
            "status": "pending_manual",
            "created_at": datetime.now().isoformat(),
            "expires_at": (datetime.now() + timedelta(days=int(days))).isoformat()
        }
        db['orders'].append(order)
        _add_or_update_client(db, client_name, client_whatsapp, client_email, order)
        save_db(db)
        return jsonify({
            "success": True,
            "manual": True,
            "message": f"✅ Pago recibido. Te enviaremos la invitación de Canva Pro a {client_email} en las próximas horas. Orden: #{order_id}"
        })
    
    # Auto-deliver account/profile
    available = [a for a in db['accounts'] if a['platform_id'] == platform_id and a['status'] == 'available']
    
    if not available:
        return jsonify({"success": False, "message": "Sin cuentas disponibles en este momento. Contáctanos por WhatsApp."})
    
    account = available[0]
    account['status'] = 'assigned'
    account['assigned_to'] = client_name
    account['assigned_email'] = client_email
    account['assigned_at'] = datetime.now().isoformat()
    account['expires_at'] = (datetime.now() + timedelta(days=int(days))).isoformat()
    
    # Reduce stock
    platform['stock'] = max(0, platform.get('stock', 1) - 1)
    
    order_id = str(uuid.uuid4())[:8].upper()
    order = {
        "id": order_id,
        "platform_id": platform_id,
        "platform_name": platform['name'],
        "client_name": client_name,
        "client_whatsapp": client_whatsapp,
        "client_email": client_email,
        "days": days,
        "price": platform['prices'].get(days, 0),
        "payment_method": method,
        "payment_code": code,
        "account_id": account['id'],
        "status": "completed",
        "created_at": datetime.now().isoformat(),
        "expires_at": account['expires_at']
    }
    db['orders'].append(order)
    
    # Register income
    db['finances']['ingresos'].append({
        "date": datetime.now().isoformat(),
        "amount": float(platform['prices'].get(days, 0)),
        "description": f"{platform['name']} - {client_name}",
        "order_id": order_id
    })
    
    _add_or_update_client(db, client_name, client_whatsapp, client_email, order)
    save_db(db)
    
    delivery_info = {}
    if platform['delivery'] == 'perfil':
        delivery_info = {
            "type": "perfil",
            "email": account.get('email', ''),
            "password": account.get('password', ''),
            "profile_name": account.get('profile_name', ''),
            "profile_pin": account.get('profile_pin', ''),
        }
    elif platform['delivery'] in ['cuenta', 'correo', 'dispositivo']:
        delivery_info = {
            "type": platform['delivery'],
            "email": account.get('email', ''),
            "password": account.get('password', ''),
            "extra": account.get('extra', '')
        }
    
    return jsonify({
        "success": True,
        "order_id": order_id,
        "delivery": delivery_info,
        "expires_at": account['expires_at'],
        "message": "✅ ¡Pago verificado! Aquí están tus credenciales."
    })

def _add_or_update_client(db, name, whatsapp, email, order):
    existing = next((c for c in db['clients'] if c.get('whatsapp') == whatsapp), None)
    if existing:
        existing['orders'].append(order['id'])
        existing['last_order'] = datetime.now().isoformat()
    else:
        db['clients'].append({
            "id": str(uuid.uuid4())[:8],
            "name": name,
            "whatsapp": whatsapp,
            "email": email,
            "orders": [order['id']],
            "created_at": datetime.now().isoformat(),
            "last_order": datetime.now().isoformat()
        })

# ─── Admin API ────────────────────────────────────────────────────────────────
@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    data = request.json
    db = load_db()
    if data.get('password') == db['admin_password']:
        session['admin'] = True
        return jsonify({"success": True})
    return jsonify({"success": False, "message": "Contraseña incorrecta"})

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin'):
            return jsonify({"success": False, "message": "No autorizado"}), 401
        return f(*args, **kwargs)
    return decorated

@app.route('/api/admin/dashboard', methods=['GET'])
@admin_required
def dashboard():
    db = load_db()
    today = datetime.now().date().isoformat()
    total_income = sum(i['amount'] for i in db['finances']['ingresos'])
    total_expenses = sum(e['amount'] for e in db['finances']['egresos'])
    today_orders = [o for o in db['orders'] if o['created_at'][:10] == today]
    expiring_soon = []
    for o in db['orders']:
        if o.get('expires_at') and o['status'] == 'completed':
            exp = datetime.fromisoformat(o['expires_at'])
            if 0 <= (exp - datetime.now()).days <= 5:
                expiring_soon.append(o)
    return jsonify({
        "total_clients": len(db['clients']),
        "total_orders": len(db['orders']),
        "today_orders": len(today_orders),
        "total_income": total_income,
        "total_expenses": total_expenses,
        "profit": total_income - total_expenses,
        "expiring_soon": len(expiring_soon),
        "platforms_count": len(db['platforms'])
    })

@app.route('/api/admin/platforms', methods=['GET','POST','PUT','DELETE'])
@admin_required
def admin_platforms():
    db = load_db()
    if request.method == 'GET':
        return jsonify(db['platforms'])
    elif request.method == 'POST':
        p = request.json
        p['id'] = p.get('id') or p['name'].lower().replace(' ','_')
        db['platforms'].append(p)
        save_db(db)
        return jsonify({"success": True})
    elif request.method == 'PUT':
        p = request.json
        for i, pl in enumerate(db['platforms']):
            if pl['id'] == p['id']:
                db['platforms'][i] = p
                break
        save_db(db)
        return jsonify({"success": True})
    elif request.method == 'DELETE':
        pid = request.args.get('id')
        db['platforms'] = [p for p in db['platforms'] if p['id'] != pid]
        save_db(db)
        return jsonify({"success": True})

@app.route('/api/admin/accounts', methods=['GET','POST','PUT','DELETE'])
@admin_required
def admin_accounts():
    db = load_db()
    if request.method == 'GET':
        platform_id = request.args.get('platform_id')
        accounts = db['accounts']
        if platform_id:
            accounts = [a for a in accounts if a['platform_id'] == platform_id]
        return jsonify(accounts)
    elif request.method == 'POST':
        acc = request.json
        acc['id'] = str(uuid.uuid4())[:8]
        acc['status'] = 'available'
        acc['created_at'] = datetime.now().isoformat()
        db['accounts'].append(acc)
        # Update stock
        for p in db['platforms']:
            if p['id'] == acc['platform_id']:
                p['stock'] = len([a for a in db['accounts'] if a['platform_id'] == acc['platform_id'] and a['status'] == 'available']) + 1
        save_db(db)
        return jsonify({"success": True, "id": acc['id']})
    elif request.method == 'PUT':
        acc = request.json
        for i, a in enumerate(db['accounts']):
            if a['id'] == acc['id']:
                db['accounts'][i] = acc
                break
        save_db(db)
        return jsonify({"success": True})
    elif request.method == 'DELETE':
        aid = request.args.get('id')
        db['accounts'] = [a for a in db['accounts'] if a['id'] != aid]
        save_db(db)
        return jsonify({"success": True})

@app.route('/api/admin/orders', methods=['GET'])
@admin_required
def admin_orders():
    db = load_db()
    return jsonify(sorted(db['orders'], key=lambda x: x['created_at'], reverse=True))

@app.route('/api/admin/clients', methods=['GET'])
@admin_required
def admin_clients():
    db = load_db()
    search = request.args.get('search', '').lower()
    clients = db['clients']
    if search:
        clients = [c for c in clients if search in c.get('name','').lower() or search in c.get('whatsapp','')]
    return jsonify(clients)

@app.route('/api/admin/finances', methods=['GET','POST'])
@admin_required
def admin_finances():
    db = load_db()
    if request.method == 'GET':
        return jsonify(db['finances'])
    elif request.method == 'POST':
        data = request.json
        if data['type'] == 'egreso':
            db['finances']['egresos'].append({
                "date": datetime.now().isoformat(),
                "amount": float(data['amount']),
                "description": data['description']
            })
        save_db(db)
        return jsonify({"success": True})

@app.route('/api/admin/export-excel', methods=['GET'])
@admin_required
def export_excel():
    db = load_db()
    wb = openpyxl.Workbook()
    
    # Clients sheet
    ws1 = wb.active
    ws1.title = "Clientes"
    ws1.append(["ID","Nombre","WhatsApp","Email","Órdenes","Fecha Registro","Último Pedido"])
    for c in db['clients']:
        ws1.append([c.get('id'),c.get('name'),c.get('whatsapp'),c.get('email'),len(c.get('orders',[])),c.get('created_at','')[:10],c.get('last_order','')[:10]])
    
    # Orders sheet
    ws2 = wb.create_sheet("Órdenes")
    ws2.append(["Orden","Plataforma","Cliente","WhatsApp","Email","Días","Precio","Método Pago","Estado","Vencimiento"])
    for o in db['orders']:
        ws2.append([o.get('id'),o.get('platform_name'),o.get('client_name'),o.get('client_whatsapp'),o.get('client_email'),o.get('days'),o.get('price'),o.get('payment_method'),o.get('status'),o.get('expires_at','')[:10] if o.get('expires_at') else ''])
    
    # Accounts sheet
    ws3 = wb.create_sheet("Cuentas")
    ws3.append(["ID","Plataforma","Email","Contraseña","Perfil","PIN","Estado","Asignado a","Vence"])
    for a in db['accounts']:
        ws3.append([a.get('id'),a.get('platform_id'),a.get('email'),a.get('password'),a.get('profile_name',''),a.get('profile_pin',''),a.get('status'),a.get('assigned_to',''),a.get('expires_at','')[:10] if a.get('expires_at') else ''])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="streamvault_export.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/admin/pending-orders', methods=['GET'])
@admin_required
def pending_orders():
    db = load_db()
    pending = [o for o in db['orders'] if o['status'] == 'pending_manual']
    return jsonify(pending)

@app.route('/api/admin/complete-order', methods=['POST'])
@admin_required
def complete_order():
    db = load_db()
    data = request.json
    order_id = data.get('order_id')
    for o in db['orders']:
        if o['id'] == order_id:
            o['status'] = 'completed'
            o['completed_at'] = datetime.now().isoformat()
            # Register income
            db['finances']['ingresos'].append({
                "date": datetime.now().isoformat(),
                "amount": float(o.get('price', 0)),
                "description": f"{o['platform_name']} - {o['client_name']}",
                "order_id": order_id
            })
            break
    save_db(db)
    return jsonify({"success": True})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
